import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from bs4 import BeautifulSoup
import requests
from inc.config import URL_XML, PATH_XLSX_FILE

def get_list_contact():
    resource = requests.get(URL_XML)
    soup = BeautifulSoup(resource.text, "xml")
    dictionary = {}
    surnames = {}
    phonerecords = []
    additional = {}
    mainnumber = {}

    for menu in soup.findAll("Menu"):
        menu_name = menu.get("Name", "Неизвестное меню")
        menu_number = menu.get("Number", "Неизвестный номер")
        for tag in menu.find_all("Unit"):
            additional[tag["Phone1"]] = tag["Name"]
            mainnumber[tag["Phone2"]] = tag["Name"]
            
            surname = tag["Name"].split()[0]
            if surname not in surnames:
                surnames[surname] = []
            surnames[surname].append(tag["Name"])

            phone1 = tag.get("Phone1", "")
            phone2 = tag.get("Phone2", "")
            phone2_str = f'📱Телефон: {phone2}' if phone2 else ""

            # Формируем строку для добавочного кода, если он есть
            phone1_str = f'☎️Добавочный код: {phone1}' if phone1 else ""

            # Объединяем строки с проверкой на наличие значений
            phone_info = "\n".join(filter(None, [phone1_str, phone2_str]))

            dictionary[tag["Name"]] = (
                f'\n👤{tag["Name"]} {(tag.get("Middle") or "")}\n'
                f'🏢{menu_name} {menu_number}\n'
                f'{phone_info}'
            )

            phonerecords.append(
                f'______________________________\n'
                f'👤ФИО: {tag["Name"]} {(tag.get("Middle") or "")}\n'
                f'🏢{menu_name} {menu_number}\n'
                f'{phone_info}'
            )

    # Сортировка phonerecords по алфавиту на основе тега NAME
    phonerecords.sort(key=lambda record: record.split('\n')[1].split(': ')[1])

    return dictionary, phonerecords, surnames, additional, mainnumber

def get_unit_contact(name):
    contacts = get_list_contact()
    phones = contacts[0]
    surnames = contacts[2]
    additional = contacts[3]
    mainnumber = contacts[4]
    error = 'Информации по запросу нет'
    nametitle = name.title()

    if name.startswith('8') and len(name) > 3:
        result = phones.get(mainnumber.get(name))
        return result if result else error
    elif name.isdigit():
        print(type(name))
        result = phones.get(additional.get(name))
        return result if result else error
    else:
        if surnames.get(nametitle) is not None:
            # Получаем список имен по фамилии
            name_list = surnames.get(nametitle)
            # Собираем все номера телефонов
            phone_numbers = [phones.get(n) for n in name_list if phones.get(n) is not None]
            if phone_numbers:
                # Форматируем вывод
                formatted_output = "\n_________________________\n".join(phone_numbers)
                return formatted_output
            else:
                return error
        else:
            result = phones.get(nametitle)
            return result if result else error

def getpagephones():
    pagephones = []
    for i in range(0, len(get_list_contact()[1]), 10):
        pagephones.append(get_list_contact()[1][i:i+10])
    return pagephones
    
def generate_xlsx():
    response = requests.get(URL_XML)
    soup = BeautifulSoup(response.content, 'xml')

    departments = {} 

    for menu in soup.find_all('Menu'):
        department_name = menu.get('Name')
        department_phone = menu.get('Number', '')

        for unit in menu.find_all('Unit'):
            unit_name = unit.get('Name')
            unit_middle = unit.get('Middle')
            internal_phone = unit.get('Phone1')
            corporate_phone = unit.get('Phone2')
            unit_email = unit.get('Email')
            unit_job_title = unit.get('JobTitle')

            contact = {
                'Name': unit_name,
                'Middle': unit_middle,
                'JobTitle': unit_job_title,
                'InternalPhone': internal_phone,
                'CorporatePhone': corporate_phone,
                'Email': unit_email,
            }

            if department_name not in departments:
                departments[department_name] = {'phone': department_phone, 'contacts': []}
            departments[department_name]['contacts'].append(contact)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Лист 1'

    # Шрифт для всей таблицы
    default_font = Font(name='Franklin Gothic Book', size=12)

    # Заголовок
    sheet.merge_cells('A1:E1')
    title_cell = sheet['A1']
    title_cell.value = 'Список телефонов сотрудников ООО "БМУ ГЭМ"'
    title_cell.font = Font(name='Franklin Gothic Book', size=12, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Заголовки столбцов
    headers = ["ФИО", "Должность", "Внутр.", "Сотовый тел.", "Эл.почта"]
    sheet.append(headers)
    for col in range(1, 6):
        cell = sheet.cell(row=2, column=col)
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def add_department_contacts(department, phone, contacts):
        department_with_phone = f"{department}, городской тел. {phone}" if phone else department
        sheet.append([department_with_phone, "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
        department_cell = sheet.cell(row=sheet.max_row, column=1)
        department_cell.font = Font(name='Franklin Gothic Book', bold=True)
        department_cell.alignment = Alignment(horizontal='center', vertical='center')

        for contact in contacts:
            sheet.append([
                f"{contact['Name']} {contact['Middle']}",
                contact['JobTitle'],
                contact['InternalPhone'],
                contact['CorporatePhone'],
                contact['Email']
            ])
            for col in range(1, 6):
                cell = sheet.cell(row=sheet.max_row, column=col)
                cell.font = default_font
                if col == 1 or col == 2 or col == 5:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    irkutsk_departments = ["АУП Иркутск", "ДПЭС", "Иркутский филиал", "Дорожная 1"]
    bratsk_departments = ["АУП Братск", "Братская Площадка", "Наладка"]

    # Город Иркутск
    sheet.append(["город Иркутск", "", "", "", ""])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
    city_cell = sheet.cell(row=sheet.max_row, column=1)
    city_cell.font = Font(name='Franklin Gothic Book', bold=True, size=14)
    city_cell.alignment = Alignment(horizontal='center', vertical='center')

    for department in irkutsk_departments:
        if department in departments:
            contacts = departments[department]['contacts']
            phone = departments[department]['phone']
            contacts.sort(key=lambda x: (x.get("Name", ""), x.get("Middle", "")))
            add_department_contacts(department, phone, contacts)

    # Город Братск
    sheet.append(["город Братск", "", "", "", ""])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
    city_cell = sheet.cell(row=sheet.max_row, column=1)
    city_cell.font = Font(name='Franklin Gothic Book', bold=True, size=14)
    city_cell.alignment = Alignment(horizontal='center', vertical='center')

    for department in bratsk_departments:
        if department in departments:
            contacts = departments[department]['contacts']
            phone = departments[department]['phone']
            contacts.sort(key=lambda x: (x.get("Name", ""), x.get("Middle", "")))
            add_department_contacts(department, phone, contacts)

    # Остальные департаменты
    remaining_departments = set(departments.keys()) - set(irkutsk_departments) - set(bratsk_departments)
    for department in remaining_departments:
        contacts = departments[department]['contacts']
        phone = departments[department]['phone']
        contacts.sort(key=lambda x: (x.get("Name", ""), x.get("Middle", "")))
        add_department_contacts(department, phone, contacts)

    # Добавление границ
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    # Установка ширины колонок вручную
    column_widths = {
        'A': 40,  # Ширина для столбца "ФИО"
        'B': 55,  # Ширина для столбца "Должность"
        'C': 10,  # Ширина для столбца "Внутр."
        'D': 20,  # Ширина для столбца "Сотовый тел."
        'E': 35   # Ширина для столбца "Эл.почта"
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Сохраняем файл
    workbook.save(PATH_XLSX_FILE)
